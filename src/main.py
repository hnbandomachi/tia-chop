from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.comments import Comment

def numOfItems(data):
    for i in range(1, len(data)):
        if data[i][0] == "TỔNG CỘNG":
            return i     

def forControl(A, B, row):
    for i in range(len(A)):
        if str(A[i]) == 'nan':
            A[i] = 0
        if str(B[i]) == 'nan':
            B[i] = 0

    if A[4] != B[4]:
        print("Ma don hang " + str(A[2]) + " bi KHAC SL dat don")
        ws['E'+str(row+7)].fill     = PatternFill("solid", fgColor="00FF0000")
        ws['E'+str(row+7)].comment  = Comment("LF: " + str(B[4]) + " \nTC: " + str(A[4]), 'Huy Le')
        ws['C'+str(row+7)].fill     = PatternFill("solid", fgColor="00FF0000")

    if A[6] != B[6]:
        print("Ma don hang " + str(A[2]) + " bi KHAC SL NCC cancel")
        ws['G'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['G'+str(row+7)].comment  = Comment("LF: " + str(B[6]) + " \nTC: " + str(A[6]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[7] != B[7]:
        print("Ma don hang " + str(A[2]) + " bi KHAC SL fail QC")
        ws['H'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['H'+str(row+7)].comment  = Comment("LF: " + str(B[7]) + " \nTC: " + str(A[7]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[10] != B[10]:
        print("Ma don hang " + str(A[2]) + " bi KHAC Don gia nhap kho")
        ws['K'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['K'+str(row+7)].comment  = Comment("LF: " + str(B[10]) + " \nTC: " + str(A[10]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[11] != B[11]:
        print("Ma don hang " + str(A[2]) + " bi KHAC Thanh tien nhap kho")
        ws['L'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['L'+str(row+7)].comment  = Comment("LF: " + str(B[11]) + " \nTC: " + str(A[11]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[12] != B[12]:
        print("Ma don hang " + str(A[2]) + " bi KHAC VAT")
        ws['M'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['M'+str(row+7)].comment  = Comment("LF: " + str(B[12]) + " \nTC: " + str(A[12]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[13] != B[13]:
        print("Ma don hang " + str(A[2]) + " bi KHAC Thanh tien thanh toan")
        ws['N'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['N'+str(row+7)].comment  = Comment("LF: " + str(B[13]) + " \nTC: " + str(A[13]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")

    if A[14] != B[14]:
        print("Ma don hang " + str(A[2]) + " bi khac thanh tien can tru")
        ws['O'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")
        ws['O'+str(row+7)].comment  = Comment("LF: " + str(B[14]) + " \nTC: " + str(A[14]), 'Huy Le')
        ws['C'+str(row+7)].fill = PatternFill("solid", fgColor="00FF0000")



def main():
    global wb, ws

    shutil.copy2(os.getcwd() + '\TC.xlsx', os.getcwd() + '\ket_qua.xlsx')
    TC_data = pd.read_excel(os.getcwd() + '\TC.xlsx', header=5).values
    LF_data = pd.read_excel(os.getcwd() + '\LF.xlsx', header=5).values

    wb = load_workbook(os.getcwd() + '\ket_qua.xlsx')
    ws = wb[wb.sheetnames[0]]

    num_TC_item = numOfItems(TC_data)    
    num_LF_item = numOfItems(LF_data)    

    # Clean data
    TC_data = TC_data[0:num_TC_item]
    LF_data = LF_data[0:num_LF_item]

    # Get SKU and PO to compare
    is_TC_reserve = True
    is_LF_reserve = True
    TC_reserves = []
    LF_reserves = []
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']   
    for i in range(num_TC_item):
        for j in range(num_LF_item):
            if TC_data[i][2] == LF_data[j][2] and TC_data[i][1] == LF_data[j][1]:      # Consider updating new approximate comparation
                a = TC_data[i][2]
                b = LF_data[j][2]
                forControl(TC_data[i][:], LF_data[j][:], i)
                is_TC_reserve = False
                break
            else:
                is_TC_reserve = True
        if True == is_TC_reserve:
            # TC_reserves.append(TC_data[i])
            for k in range(len(cols)):
                ws[cols[k] + str(i+7)]        = TC_data[i][k]
                ws[cols[k] + str(i+7)].fill   = PatternFill("solid", fgColor="09EA69")

    # for TC_reserve in TC_reserves:
    #     for i in range(len(cols)):
    #         ws[cols[i] + str(num_TC_item+7)]        = TC_reserve[i]
    #         ws[cols[i] + str(num_TC_item+7)].fill   = PatternFill("solid", fgColor="09EA69")


    wb.save(os.getcwd() + '\ket_qua.xlsx')
    print("Hello users of Ngoc-Bui, this program is for you... enjoy it :)))")

if __name__ == "__main__":
    main()